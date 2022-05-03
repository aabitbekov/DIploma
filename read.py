import os

import numpy
import numpy as np
import openpyxl
from pathlib import Path


def save_uploadedfile(uploadedfile):
    with open(os.path.join(uploadedfile.name), "wb") as f:
        f.write(uploadedfile.getbuffer())
    return uploadedfile.name



def koef_prymyx_zatrat(main_matrix, val_price):
    koef_prymyx_zatrat = []
    row_koef_prymyx_zatrat = []
    index = 0
    for array in main_matrix:
        for element in array:
            row_koef_prymyx_zatrat.append(round(element/val_price[index], 4))
            index += 1
        koef_prymyx_zatrat.append(row_koef_prymyx_zatrat)
        row_koef_prymyx_zatrat = []
        index = 0
    return koef_prymyx_zatrat





def readDocument(path):
    xlsx_file = Path(path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    return sheet


def readMainMatrix(sheet):
    # Read Matrix
    main_matrix, raw = [], []
    for row in sheet.iter_rows(2, sheet.max_row-5):
        for index in range(1,sheet.max_column-3):
            raw.append(row[index].value)
        main_matrix.append(raw)
        raw = []
    return main_matrix

def readByCol(sheet):
    # Read By Col
    col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy = [], [], [], [], [], []
    for column in sheet.iter_cols(2, sheet.max_column):
        col_names.append(column[0].value)
        itogo_by_row.append(column[sheet.max_row - 5].value)
        dob_st.append(column[sheet.max_row - 4].value)
        val_price_by_raw.append(column[sheet.max_row - 3].value)
        trud.append(column[sheet.max_row - 2].value)
        fondy.append(column[sheet.max_row - 1].value)
    return col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy

def readByRow(sheet):
    # Read By Row
    row_names, itogo_by_col, end_pruducts, val_price = [], [], [], []
    for row in sheet.iter_rows(2, sheet.max_row):
        row_names.append(row[0].value)
        itogo_by_col.append(row[sheet.max_column - 3].value)
        end_pruducts.append(row[sheet.max_column - 2].value)
        val_price.append(row[sheet.max_column - 1].value)
    return row_names, itogo_by_col, end_pruducts, val_price


def getInverseMatrix(main_matrix):
    inverse_matrix, identity_matrix = [], np.identity(len(main_matrix))
    b = numpy.linalg.inv(np.subtract(identity_matrix, main_matrix))
    for i in range(len(b)):
        for j in range(len(b)):
            b[i][j] = round(b[i][j], 3)

    return b


if __name__ == '__main__':
    # main_matrix, row_names, itogo_by_col, end_pruducts, val_price, col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy \
    sheet = readDocument('data.xlsx')
    main_matrix = readMainMatrix(sheet)
    row_names, itogo_by_col, end_pruducts, val_price = readByRow(sheet)
    col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy = readByCol(sheet)

    print(list(filter(None, trud)), val_price_by_raw)







#
# def readDocument(path):
#     xlsx_file = Path(path)
#     wb_obj = openpyxl.load_workbook(xlsx_file)
#     sheet = wb_obj.active
#     main_matrix = []
#     var_row, total_count , kon_product, val_price = [],[], [], []
#     outiter, inter = 0, 0
#     for row in sheet.iter_rows(max_col=sheet.max_column, max_row=sheet.max_row):
#         if outiter == 0:
#             pass
#         else:
#             inter = 0
#             for cell in row:
#                 if inter == 0:
#                     pass
#                 else:
#                     if inter == sheet.max_row - 1:
#                         total_count.append(cell.value)
#                         # print(total_count)
#                     elif inter == sheet.max_row:
#                         kon_product.append(cell.value)
#                         # print(kon_product)
#                     elif inter == sheet.max_row + 1:
#                         val_price.append(cell.value)
#                         # print(val_price)
#                     else:
#                         var_row.append(cell.value)
#                 inter += 1
#         if var_row != []:
#             main_matrix.append(var_row)
#             var_row = []
#         outiter += 1
#
#     # print(main_matrix)
#     col_names = []
#     for column in sheet.iter_cols(1, sheet.max_column):
#         col_names.append(column[0].value)
#     # print(col_names)
#     row_names = []
#     for row in sheet.iter_rows(1, sheet.max_row):
#         row_names.append(row[0].value)
#     # print(row_names)
#     return main_matrix, col_names, row_names, total_count, kon_product, val_price



